# -*- coding: utf-8 -*-
"""
Created on Mon Apr 19 11:23:52 2021

@author: sungmin.hwang
"""

from bs4 import BeautifulSoup
from selenium import webdriver
import time
import sys


query_txt = input("크롤링할 키워드=")
f_name = input("txt로 저정할 경로=")
fc_name = input("csv로 저정할 경로=")
fx_name = input("xls로 저정할 경로=")


path=("D:\\temp\\chromedriver.exe")
#driver = webdriver.chrome(path)
driver = webdriver.Chrome(path)
driver.get("https://korean.visitkorea.or.kr")

time.sleep(2)
driver.find_element_by_xpath("/html/body/div[5]/div/div/div/button").click()


#search_button = driver.find_element_by_xpath("/html/body/header/div[5]/div/div/div//button")





element = driver.find_element_by_id("inp_search")

element.send_keys(query_txt)
time.sleep(1)
driver.find_element_by_link_text("검색").click()


time.sleep(1)

html = driver.page_source

soup = BeautifulSoup(html,'html.parser')

content_list = soup.find('ul', class_= 'list_thumType type1')

#print(content_list)

#학습목표 : 특정 항목을 분리해서 추출하기
contents = content_list.find('div','tit').get_text()
print('내용:',contents.strip())

tag = content_list.find('p','tag_type').get_text()
print('태그',tag.strip())
print('\n')

#각 항목별로 분리하여 추출하고 변수에 할당하기

no=1
no2=[]
contents2=[]
tags2=[]

for i in content_list:
    try:
        contents = i.find('div','tit').get_text()
        contents2.append(contents)
        print('내용',contents.strip())
    
        no2.append(no)
        print('번호',no)    
        
        tag = i.find('p','tag_type').get_text()
        tags2.append(tag)
        print('태그',tag.strip())
        print('\n')
        no += 1
    except :
        pass


#csv 파일 형식 저장
import pandas as pd

korea_trip = pd.DataFrame()
korea_trip['번호'] = no2
korea_trip['내용'] = contents2
korea_trip['태그'] = tags2

korea_trip.to_csv(fc_name,encoding='utf-8-sig')
print("csv 파일 저장 경로= %s"%fc_name)



#엑셀 형식으로 저장
import xlwt
korea_trip.to_excel(fx_name)
print("xls 파일 저장경로=%s"%fx_name)


f = open(f_name,'a',encoding='UTF-8')
f.write(str(contents2))
f.write(str(tags2))
f.write(str('\n'))
f.close()

import openpyxl
wb = openpyxl.Workbook()
wb.save("d:\\temp\\test1.xlsx")

import openpyxl
wb = openpyxl.Workbook()

sheet_1 = wb.active
sheet_2 = wb.create_sheet("매출현황")

sheet_1.title='총매출현황'

wb.save("d:\\temp\\test2.xlsx")


import openpyxl
wb = openpyxl.load_workbook('d:\\temp\\test2.xlsx')
sheet_1 = wb['총매출현황']
sheet_1['A1'] = '첫번째 cell'
sheet_1['A2'] = '두번째 cell'

wb.save("d:\\temp\\test2.xlsx")

